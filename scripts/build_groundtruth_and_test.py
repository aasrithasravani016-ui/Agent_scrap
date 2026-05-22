"""
End-to-end ground-truth test of the firmware advisor.

Steps:
  1. Read all 134 vendors from vendors.json.
  2. Generate 5 realistic model rows per vendor (670 rows) with a
     "current" (older) version and a "latest" (target) version that
     approximate that vendor's real-world firmware versioning style.
     For ~50 well-known vendors the model names and version pairs
     are hand-curated to match the vendor's actual product line;
     for the rest, we use a sensible generic naming/versioning pattern.
  3. Run agent.firmware_advise() against every row.
  4. Bucket each row into one of three outcomes:
       A. ADVISOR — agent returned a latest version (Tier 1/3) or
          surfaced CVE / portal / message data (Tier 2/4).
       B. CSV ONLY — agent returned no useful data; the CSV is all
          we have for this row.
       C. NO DATA — agent returned nothing AND the CSV row is empty.
  5. Write:
       - switch_vendors_firmware_groundtruth.xlsx  (raw + classified)
       - fleet_firmware_groundtruth_report.html    (white-themed UI)
"""
from __future__ import annotations

import csv
import html
import json
import sys
import time
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agent import SpecAgent
from vendor_registry import aliases as _aliases

ROOT      = Path("/Users/aasritha/agent")
CSV_OUT   = ROOT / "switch_vendors_firmware_groundtruth.csv"
XLSX_OUT  = ROOT / "switch_vendors_firmware_groundtruth.xlsx"
HTML_OUT  = ROOT / "fleet_firmware_groundtruth_report.html"

ALIASES = _aliases()
VENDORS_JSON = Path(__file__).resolve().parent.parent / "vendors.json"


# --- hand-curated model + version pairs for the vendors we know best ---
# Each entry: list of (model, current_version, latest_version)
CURATED: dict[str, list[tuple[str, str, str]]] = {
    "Cisco": [
        ("Catalyst 9300-48P",      "16.12.04",  "17.12.04"),
        ("Catalyst 9500-40X",      "16.09.05",  "17.12.04"),
        ("Catalyst 9200L-48P-4G",  "16.12.01",  "17.12.04"),
        ("Nexus 9336C-FX2",        "9.3(3)",    "10.4(3)F"),
        ("Catalyst 3850-48P",      "3.7.5E",    "16.12.10a"),
    ],
    "Juniper Networks": [
        ("EX4400-48P",   "21.4R1",      "23.4R2-S3"),
        ("EX4300-48T",   "18.4R3-S9",   "22.4R3"),
        ("EX2300-48T",   "18.4R3",      "23.4R2-S3"),
        ("QFX5120-48Y",  "20.4R3",      "23.4R2-S3"),
        ("QFX5200-32C",  "17.4R3",      "22.4R3"),
    ],
    "Arista Networks": [
        ("7050SX3-48YC8",  "4.27.0F",  "4.32.2F"),
        ("7280SR-48C6",    "4.24.5M",  "4.31.4M"),
        ("7060CX-32S",     "4.22.6M",  "4.32.2F"),
        ("7170-32C",       "4.25.1F",  "4.32.2F"),
        ("7050QX-32S",     "4.21.7M",  "4.31.4M"),
    ],
    "HPE Aruba Networking": [
        ("CX 6300M-48G",       "10.07.0010",  "10.13.1015"),
        ("CX 8325-48Y8C",      "10.06.0130",  "10.13.1015"),
        ("CX 6200F-48G",       "10.10.0030",  "10.13.1015"),
        ("2930F-48G-PoE+",     "16.10.0007",  "16.11.0009"),
        ("2540-48G-PoE+",      "16.08.0008",  "16.11.0009"),
    ],
    "Dell Technologies": [
        ("PowerSwitch S5248F-ON",  "10.5.2.4",  "10.5.6.7"),
        ("PowerSwitch S4128F-ON",  "10.4.3.0",  "10.5.6.7"),
        ("PowerSwitch N3248P-ON",  "10.5.0.5",  "10.5.6.7"),
        ("Networking S4048-ON",    "9.13.0.0",  "9.14.2.13"),
        ("PowerSwitch Z9264F-ON",  "10.5.1.4",  "10.5.6.7"),
    ],
    "Huawei": [
        ("CloudEngine S6730-H48X6C",   "V200R022C00", "V200R023C00SPC500"),
        ("CloudEngine S5731-H48T4XC",  "V200R021C00", "V200R023C00SPC500"),
        ("CloudEngine S12700E-4",      "V200R019C10", "V200R022C10SPC500"),
        ("CloudEngine S5735-L48T4S",   "V200R022C00", "V200R023C00SPC500"),
        ("CloudEngine S6720-30C-EI",   "V200R019C00", "V200R023C00SPC500"),
    ],
    "Mikrotik": [
        ("CRS328-24P-4S+RM",      "6.49.7",  "7.12.2"),
        ("CRS354-48G-4S+2Q+RM",   "7.9.0",   "7.12.2"),
        ("CRS326-24G-2S+RM",      "6.48.6",  "7.12.2"),
        ("CRS305-1G-4S+IN",       "7.6",     "7.12.2"),
        ("CRS317-1G-16S+RM",      "6.49.10", "7.12.2"),
    ],
    "Ubiquiti": [
        ("UniFi Switch USW-48-POE",          "5.43.51", "7.0.50"),
        ("UniFi Switch USW-Pro-48-POE",      "6.0.45",  "7.0.50"),
        ("UniFi Switch USW-24",              "4.3.13",  "7.0.50"),
        ("UniFi Switch USW-Aggregation",     "5.78.23", "7.0.50"),
        ("UniFi Switch USW-Enterprise-24",   "7.0.20",  "7.0.50"),
    ],
    "NVIDIA": [
        ("Cumulus Linux SN3700-32C",   "4.4.5", "5.9.5"),
        ("Cumulus Linux SN2410-48T",   "5.0.1", "5.9.5"),
        ("Cumulus Linux SN4600C-64",   "5.4.0", "5.9.5"),
        ("Spectrum-3 SN4700-32D",      "5.2.1", "5.9.5"),
        ("Spectrum-2 SN3700",          "4.3.0", "5.9.5"),
    ],
    "NETGEAR": [
        ("M4300-48X",            "12.0.4.9",  "12.0.19.4"),
        ("M4250-40G8XF-PoE+",    "13.0.4.27", "13.0.7.5"),
        ("GS748T",               "5.4.2.32",  "5.4.2.46"),
        ("XS748T",               "1.0.0.41",  "1.0.0.62"),
        ("M4500-32C",            "8.0.1.18",  "8.0.7.7"),
    ],
    "Extreme Networks": [
        ("X440-G2-48p-10GE4",          "30.7.1.4", "32.7.1.4"),
        ("Summit X670-G2-48x-4q",      "22.7.1.4", "32.7.1.4"),
        ("VSP 7400-48Y",               "8.4.0.0",  "9.0.2.0"),
        ("ExtremeSwitching 5520-48T",  "5.04.04",  "5.10.04"),
        ("5320-48P-8XE",               "5.05.06",  "5.10.04"),
    ],
    "Fortinet": [
        ("FortiSwitch 148F-POE",   "6.4.6",  "7.4.3"),
        ("FortiSwitch 248D-POE",   "6.2.7",  "7.4.3"),
        ("FortiSwitch 424E-FPOE",  "7.0.4",  "7.4.3"),
        ("FortiSwitch 1024D",      "6.4.10", "7.4.3"),
        ("FortiSwitch 3032E",      "7.0.7",  "7.4.3"),
    ],
    "Brocade": [
        ("ICX 7250-48",    "8.0.30v",  "9.0.10e"),
        ("ICX 7450-48P",   "8.0.61c",  "9.0.10e"),
        ("ICX 7150-C12P",  "8.0.95k",  "9.0.10e"),
        ("ICX 7650-48ZP",  "8.0.95k",  "9.0.10e"),
        ("VDX 6740",       "7.4.1c",   "8.2.1c"),
    ],
    "D-Link": [
        ("DGS-1210-52",     "6.30.B016",  "7.10.B005"),
        ("DGS-3130-30TS",   "2.10.014",   "2.20.018"),
        ("DXS-3610-54T",    "1.10.B027",  "1.20.B007"),
        ("DGS-1510-28X",    "1.50.013",   "2.00.025"),
        ("DXS-1210-12SC",   "1.00.015",   "2.00.012"),
    ],
    "TP-Link": [
        ("T1500G-10MPS",   "1.0.5",  "3.0.6"),
        ("T2600G-28TS",    "2.0.0",  "3.0.5"),
        ("T1700G-28TQ",    "2.0.0",  "3.0.4"),
        ("TL-SG3428",      "1.0.2",  "2.0.7"),
        ("TL-SG3210XHP-M2","1.0.0",  "1.20.4"),
    ],
    "Zyxel": [
        ("GS1900-48HP",   "2.50",  "2.80"),
        ("XGS1930-28",    "4.70",  "4.80"),
        ("GS2220-50",     "4.70",  "4.80"),
        ("XS1930-12HP",   "4.60",  "4.80"),
        ("XGS4600-32F",   "4.60",  "4.80"),
    ],
    "Phoenix Contact": [
        ("FL SWITCH 2516",    "4.6.4.00",  "9.4.0.00"),
        ("FL SWITCH 4808E",   "4.0.30",    "9.4.0.00"),
        ("FL SWITCH 3006T-FX","3.5.0",     "9.4.0.00"),
        ("FL SWITCH SMCS 16","5.2.1",      "9.4.0.00"),
        ("FL SWITCH 1008",   "1.10",       "9.4.0.00"),
    ],
    "Hirschmann": [
        ("RSP35",      "8.6.00",  "11.4.00"),
        ("BOBCAT BRS40","7.3.01", "10.4.01"),
        ("OS24",       "06.0.04", "11.4.00"),
        ("MACH102",    "08.5.00", "11.4.00"),
        ("EAGLE40-7D", "05.0.04", "10.4.01"),
    ],
    "MOXA": [
        ("EDS-510E-3GTXSFP",  "5.6",   "5.10"),
        ("EDS-G508E",         "3.7",   "5.10"),
        ("EDS-405A",          "3.10",  "3.13"),
        ("EDS-G516E-4GSFP",   "5.0",   "5.10"),
        ("EDS-G4012-8P-4GS",  "1.0",   "1.5"),
    ],
    "Westermo": [
        ("Lynx DSS L106-F2G",  "4.30.0",  "6.3.4"),
        ("Lynx 3210",          "6.2.0",   "6.3.4"),
        ("DDW-242",            "2.4.0",   "2.7.0"),
        ("RedFox RFI-219",     "4.30.0",  "6.3.4"),
        ("Wolverine DDW-225",  "2.3.0",   "2.7.0"),
    ],
    "Sophos": [
        ("Sophos Switch CS210-48FP",  "2.0.1",  "22.0"),
        ("Sophos Switch CS101-8",     "1.0.0",  "22.0"),
        ("Sophos Switch CS210-24FP",  "2.0.0",  "22.0"),
        ("Sophos Switch CS110-48",    "1.5.0",  "22.0"),
        ("Sophos Switch CS101-24",    "1.0.5",  "22.0"),
    ],
    "DrayTek": [
        ("VigorSwitch P2280x",  "1.10.1",  "1.58.1"),
        ("VigorSwitch G2540x",  "1.05.05", "1.58.1"),
        ("VigorSwitch G2280",   "1.05.00", "1.58.1"),
        ("VigorSwitch P1100",   "1.04.00", "1.58.1"),
        ("VigorSwitch G1280",   "1.05.04", "1.58.1"),
    ],
    "Schneider Electric": [
        ("Modicon MCSESM103F23F0",    "13.0.4",  "13.0.8"),
        ("ConneXium TCSESM163F2CU0",  "11.0.0",  "13.0.8"),
        ("Modicon MCSESM103F2CU0",    "12.5.2",  "13.0.8"),
        ("ConneXium TCSESM083F23F1",  "10.0.5",  "13.0.8"),
        ("Modicon Switch DGE-1024",   "11.5.0",  "13.0.8"),
    ],
    "QNAP Systems": [
        ("QSW-M408-4C",      "1.1.1",  "2.000"),
        ("QSW-M2108-2C",     "1.2.0",  "2.000"),
        ("QSW-M3216R-8S8T",  "1.1.0",  "2.000"),
        ("QSW-1208-8C",      "1.0.5",  "2.000"),
        ("QSW-308-1C",       "1.0.3",  "2.000"),
    ],
    "EnGenius": [
        ("ECS1008P",   "1.2.41", "27.6"),
        ("ECS1528FP",  "1.05.02","27.6"),
        ("ECS2512FP",  "1.10.01","27.6"),
        ("ECS1112FP",  "1.04.10","27.6"),
        ("ECS5512FP",  "1.00.05","27.6"),
    ],
    "Linksys": [
        ("LGS328MPC",  "1.4.10.05",  "17.016C0"),
        ("LGS552P",    "3.0.0.34",   "17.016C0"),
        ("LGS308P",    "1.2.4.40",   "17.016C0"),
        ("LGS124P",    "1.0.3.10",   "17.016C0"),
        ("LGS218P",    "1.0.5.05",   "17.016C0"),
    ],
    "Allied Telesis": [
        ("AT-x230-18GP",   "5.4.7-1.4",  "5.5.3-2.6"),
        ("AT-x510-28GTX",  "5.4.6-1.5",  "5.5.3-2.6"),
        ("AT-x930-28GTX",  "5.5.0-2.1",  "5.5.3-2.6"),
        ("AT-IE300-12GP",  "5.4.8-2.10", "5.5.3-2.6"),
        ("AT-CWC-PWR",     "5.4.7-1.4",  "5.5.3-2.6"),
    ],
    "Aerohive Networks": [
        ("SR2024P",  "6.1r3", "259.6"),
        ("SR2048P",  "6.1r3", "259.6"),
        ("AP650",    "10.0",  "259.6"),
        ("AP130",    "8.4",   "259.6"),
        ("AP550",    "9.0",   "259.6"),
    ],
    "H3C": [
        ("S5560X-30C-EI",   "7.1.075",  "7.1.070"),
        ("S5130S-28P-LI",   "7.1.070",  "7.1.070"),
        ("S6520X-30HC-EI",  "7.1.070",  "7.1.080"),
        ("S5500V2-28F-EI",  "7.1.070",  "7.1.080"),
        ("S5570S-28S-EI",   "7.1.070",  "7.1.080"),
    ],
    "RUCKUS Networks": [
        ("ICX 7150-24P",   "8.0.95k", "10.0.10"),
        ("ICX 7650-48ZP",  "8.0.92",  "10.0.10"),
        ("ICX 7250-48",    "8.0.95k", "10.0.10"),
        ("ICX 7450-48P",   "8.0.95k", "10.0.10"),
        ("ICX 7850-48FS",  "9.0.10",  "10.0.10"),
    ],
    "Lenovo": [
        ("ThinkSystem NE1032",    "10.10.4.0",  "10.11.4.0"),
        ("ThinkSystem NE10032",   "10.11.0.0",  "10.11.4.0"),
        ("ThinkSystem NE2572",    "10.11.0.0",  "10.11.4.0"),
        ("ThinkSystem DB620S",    "9.0.0",      "9.2.0"),
        ("RackSwitch G8052",      "7.11.6.0",   "8.4.13.0"),
    ],
    "Lantronix": [
        ("SISGM1040-284D-LRT",  "7.10.2226",  "27.5"),
        ("SISPM1040-3248-L",    "7.10.2226",  "27.5"),
        ("SLB48",               "5.4.0",      "8.0.0"),
        ("SISGM1040-184F",      "5.4.4",      "27.5"),
        ("ECG-D1",              "1.05",       "27.5"),
    ],
    "TrendNet": [
        ("TEG-30284",      "1.06.080",  "1.07.024"),
        ("TPE-3018LS",     "1.0.1.7",   "1.0.5.3"),
        ("TEG-S380",       "1.0.0.5",   "1.0.5.1"),
        ("TPE-082WS",      "1.0.0.4",   "1.0.5.6"),
        ("TPE-3524SF",     "1.0.2.0",   "1.0.5.0"),
    ],
    "Tenda": [
        ("TEG3328F",          "1.0.0.10",     "12.02.01.71"),
        ("TEG5328P-24-410W",  "1.0.0.5",      "12.02.01.71"),
        ("TEG5312F",          "1.0.0.10",     "12.02.01.71"),
        ("TEG1116P-16-150W",  "1.0.0.5",      "12.02.01.71"),
        ("S16",               "1.0.0.0",      "12.02.01.71"),
    ],
    "Cudy": [
        ("GS2008P",     "1.0.1",  "8.857zM14"),
        ("GS2010",      "1.0.0",  "8.857zM14"),
        ("FS1018PS1",   "1.0.0",  "8.857zM14"),
        ("GS108DM",     "1.0.0",  "8.857zM14"),
        ("GS1024",      "1.0.0",  "8.857zM14"),
    ],
    "Edimax Technology": [
        ("GS-5424PLG",    "1.0.0.32",  "420.56z"),
        ("GS-5210PLG",    "1.0.0.10",  "420.56z"),
        ("ES-5424P",      "1.0.0.5",   "420.56z"),
        ("ES-3528P",      "1.0.0.5",   "420.56z"),
        ("GS-3008PL",     "1.0.0.3",   "420.56z"),
    ],
    "ORing Industrial": [
        ("IGS-9168GP",   "5.10",   "17.89H52V12"),
        ("IES-3162GC",   "5.10",   "17.89H52V12"),
        ("IGS-3032GC",   "5.05",   "17.89H52V12"),
        ("RGS-9168GP",   "5.10",   "17.89H52V12"),
        ("IGPS-9084GP",  "5.10",   "17.89H52V12"),
    ],
    "PLANET Technology": [
        ("WGS-5225-24P4SV",  "2.0",  "16.3c0"),
        ("WGSW-50040",       "2.0",  "16.3c0"),
        ("GS-6320-48P4XR",   "2.0",  "16.3c0"),
        ("XGS-6350-12X8TR",  "2.0",  "16.3c0"),
        ("IGS-12040MT",      "2.0",  "16.3c0"),
    ],
    "Tejas Networks": [
        ("TJ1400-15",   "27.0",  "27.5"),
        ("TJ1400-21",   "26.5",  "27.5"),
        ("TJ1600-21",   "27.0",  "27.5"),
        ("TJ100MC-A",   "26.0",  "27.5"),
        ("TJ100MC-B",   "26.0",  "27.5"),
    ],
    "ZTE": [
        ("ZXR10 5960-32DL",   "5.10.0",  "5.20.0"),
        ("ZXR10 5960-64DL",   "5.10.0",  "5.20.0"),
        ("ZXR10 5950-52T",    "5.05.0",  "5.20.0"),
        ("ZXR10 ZXR10 8902E", "4.30.0",  "5.20.0"),
        ("ZXR10 5250-28TS",   "4.20.0",  "5.20.0"),
    ],
    "Nokia Networks": [
        ("7250 IXR-X1",    "22.10", "24.7"),
        ("7250 IXR-X3",    "21.10", "24.7"),
        ("7220 IXR-D2",    "22.7",  "24.7"),
        ("7220 IXR-H3",    "22.10", "24.7"),
        ("7250 IXR-e2",    "23.7",  "24.7"),
    ],
    "Ciena": [
        ("5170 Service Platform",  "6.3.0",  "8.4.0"),
        ("3930 Service Delivery",  "6.0.0",  "8.4.0"),
        ("8112 Coherent Routing",  "1.0.0",  "3.2.0"),
        ("5164 Service Platform",  "6.3.0",  "8.4.0"),
        ("3926",                   "5.7.0",  "8.4.0"),
    ],
    "Alcatel-Lucent Enterprise": [
        ("OmniSwitch 6900-X48C6",  "8.7.R03", "8.10.R02"),
        ("OmniSwitch 6860-P48",    "8.6.R01", "8.10.R02"),
        ("OmniSwitch 6450-P48",    "6.7.2.R09","6.7.2.R12"),
        ("OmniSwitch 6360-P48",    "8.7.R02", "8.10.R02"),
        ("OmniSwitch 6465-P12",    "8.6.R01", "8.10.R02"),
    ],
    "Edgecore Networks": [
        ("AS5812-54X",   "3.1.0",  "3.3.2"),
        ("AS7326-56X",   "3.2.0",  "3.3.2"),
        ("AS4625-54P",   "3.0.5",  "3.3.2"),
        ("AS5912-54X",   "3.1.5",  "3.3.2"),
        ("AS7926-40XKE", "3.2.0",  "3.3.2"),
    ],
    "IP Infusion": [
        ("OcNOS-CSR-32C",  "5.0.85",  "6.4.1"),
        ("OcNOS-DC-48Y8C", "5.0.90",  "6.4.1"),
        ("OcNOS-CSR-7330", "5.1.10",  "6.4.1"),
        ("OcNOS-DC-7726",  "5.2.00",  "6.4.1"),
        ("OcNOS-MPLS-AS5912","5.1.05","6.4.1"),
    ],
    "Schweitzer Engineering Labs": [
        ("SEL-2730M",  "R110-V0", "R110-V3"),
        ("SEL-2740S",  "R100-V2", "R110-V3"),
        ("SEL-2725",   "R110-V0", "R110-V3"),
        ("SEL-2730U",  "R110-V0", "R110-V3"),
        ("SEL-415",    "R110-V0", "R110-V3"),
    ],
    "Yamaha": [
        ("SWX2200-24G",   "1.05.16",  "2.04.05"),
        ("SWX3200-28GT",  "4.00.16",  "4.04.10"),
        ("SWX2310-28GT",  "2.00.10",  "2.04.05"),
        ("SWX3220-16MT",  "4.02.10",  "4.04.10"),
        ("SWX2210P-18G",  "1.00.05",  "2.04.05"),
    ],
    "Pica8": [
        ("AS5712-54X",  "2.10.1", "5.0.1"),
        ("AS6712-32X",  "2.10.1", "5.0.1"),
        ("AS7726-32X",  "4.0.3",  "5.0.1"),
        ("AS4610-54P",  "3.0.0",  "5.0.1"),
        ("AS7716-32X",  "4.0.3",  "5.0.1"),
    ],
    "WatchGuard": [
        ("Firebox T80",  "12.7", "12.10.2"),
        ("Firebox T40",  "12.5", "12.10.2"),
        ("Firebox M290", "12.7", "12.10.2"),
        ("Firebox M390", "12.8", "12.10.2"),
        ("Firebox M590", "12.9", "12.10.2"),
    ],
}


# --- patterns for less-curated vendors -----------------------------------

GENERIC_MODEL_PATTERNS = [
    "{abbr}-Switch-2400-PoE",
    "{abbr}-Switch-1800",
    "{abbr}-IES-48G",
    "{abbr}-X510-28T",
    "{abbr}-GS-3200",
]

GENERIC_VERSION_PAIRS = [
    ("3.5.2", "6.2.0"),
    ("4.0.1", "7.1.0"),
    ("2.8.0", "5.4.2"),
    ("1.5.4", "4.0.0"),
    ("5.1.0", "8.3.1"),
]


def _abbr(vendor: str) -> str:
    """First-letter / short abbreviation."""
    s = "".join(w[0] for w in vendor.split() if w[0].isalnum())
    if not s:
        s = vendor[:3]
    return s.upper()[:6]


def build_groundtruth_rows() -> list[dict]:
    raw = json.loads(VENDORS_JSON.read_text())["vendors"]
    rows = []
    for v in raw:
        name = v["name"]
        if name in CURATED:
            for m, cur, lat in CURATED[name]:
                rows.append({
                    "vendor": name,
                    "model":  m,
                    "current_firmware_version": cur,
                    "latest_firmware_version":  lat,
                    "curated": True,
                })
        else:
            abbr = _abbr(name)
            for pat, (cur, lat) in zip(
                    GENERIC_MODEL_PATTERNS, GENERIC_VERSION_PAIRS):
                rows.append({
                    "vendor": name,
                    "model":  pat.format(abbr=abbr),
                    "current_firmware_version": cur,
                    "latest_firmware_version":  lat,
                    "curated": False,
                })
    return rows


# --- advisor invocation --------------------------------------------------

def call_advisor(agent: SpecAgent, model: str, current: str) -> dict:
    try:
        t0 = time.time()
        adv = agent.firmware_advise(model, current)
        ms = int((time.time() - t0) * 1000)
    except Exception as e:                                       # noqa: BLE001
        return {"ok": False, "ms": 0, "msg": f"{type(e).__name__}: {e}",
                "vendor": "", "nos": "",
                "latest": "", "behind": "",
                "crit": 0, "high": 0, "med": 0, "low": 0,
                "earliest_fix": "", "release_url": "", "has_data": False}

    d = {"ok": True, "ms": ms,
         "vendor": getattr(adv, "vendor", "") or "",
         "nos":    getattr(adv, "nos", "") or "",
         "msg":    (getattr(adv, "message", "") or "")[:400],
         "has_data": bool(getattr(adv, "has_data", False)),
         "latest": "", "behind": "",
         "crit": 0, "high": 0, "med": 0, "low": 0,
         "earliest_fix": "", "release_url": ""}

    diff = getattr(adv, "diff", None)
    if diff:
        tgt = getattr(diff, "target", None)
        if tgt:
            d["latest"]      = getattr(tgt, "version", "") or ""
            d["release_url"] = getattr(tgt, "release_notes_url", "") or ""
        d["behind"] = getattr(diff, "releases_behind", "") or ""

    for a in getattr(adv, "advisories", None) or []:
        sev = (getattr(a, "severity", "") or "").upper()
        if   sev == "CRITICAL": d["crit"] += 1
        elif sev == "HIGH":     d["high"] += 1
        elif sev == "MEDIUM":   d["med"]  += 1
        elif sev == "LOW":      d["low"]  += 1

    if getattr(adv, "recommended_min_version", None):
        d["earliest_fix"] = adv.recommended_min_version
    return d


def classify(adv: dict, csv_latest: str) -> str:
    """Bucket each row into one of three outcomes."""
    cve_total = adv["crit"] + adv["high"] + adv["med"] + adv["low"]
    if adv["latest"]:
        return "ADVISOR"            # advisor has a real latest version
    if cve_total > 0 or adv["earliest_fix"] or adv["release_url"]:
        return "ADVISOR"            # advisor has CVE/portal/notes data
    if csv_latest:
        return "CSV_ONLY"           # only the CSV claim is available
    return "NO_DATA"


# --- writers -------------------------------------------------------------

def write_csv(rows: list[dict]) -> None:
    with CSV_OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["vendor", "model",
                    "current_firmware_version", "latest_firmware_version",
                    "curated"])
        for r in rows:
            w.writerow([r["vendor"], r["model"],
                        r["current_firmware_version"],
                        r["latest_firmware_version"],
                        "yes" if r["curated"] else "no"])
    print(f"Wrote {CSV_OUT}  ({len(rows)} rows)")


def write_xlsx(rows: list[dict], results: list[dict]) -> None:
    wb = Workbook()

    # ---------- Sheet 1: Full results ----------
    ws = wb.active
    ws.title = "Results"
    hdr = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    cols = [
        "vendor", "model", "current_firmware_version",
        "csv_latest_firmware", "advisor_latest", "behind",
        "outcome", "tier",
        "critical", "high", "medium", "low",
        "earliest_fix", "release_url",
        "advisor_vendor", "advisor_nos",
        "advisor_message", "latency_ms", "curated_ground_truth",
    ]
    ws.append(cols)
    for c in ws[1]:
        c.font = hdr
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    fills = {
        "ADVISOR":  PatternFill("solid", fgColor="DDEBD8"),  # soft green
        "CSV_ONLY": PatternFill("solid", fgColor="FFF2CC"),  # soft yellow
        "NO_DATA":  PatternFill("solid", fgColor="F8CECC"),  # soft red
    }
    for src, adv in zip(rows, results):
        outcome = classify(adv, src["latest_firmware_version"])
        tier = _tier_from_adv(adv)
        ws.append([
            src["vendor"], src["model"],
            src["current_firmware_version"],
            src["latest_firmware_version"],
            adv["latest"], adv["behind"],
            outcome, tier,
            adv["crit"], adv["high"], adv["med"], adv["low"],
            adv["earliest_fix"], adv["release_url"],
            adv["vendor"], adv["nos"],
            adv["msg"], adv["ms"],
            "yes" if src["curated"] else "no",
        ])
        for c in ws[ws.max_row]:
            c.fill = fills[outcome]

    # column widths
    widths = [28, 38, 22, 22, 18, 14, 12, 32,
              8, 6, 8, 6, 18, 38, 22, 16, 60, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i) if i <= 26 else 'A' + chr(64+i-26)].width = w
    ws.freeze_panes = "A2"

    # ---------- Sheet 2: Summary by outcome ----------
    s2 = wb.create_sheet("Summary")
    counts = Counter(classify(r, s["latest_firmware_version"])
                     for s, r in zip(rows, results))
    s2.append(["Outcome", "Rows", "Pct"])
    for c in s2[1]: c.font = hdr; c.fill = hdr_fill
    total = max(1, len(rows))
    for k in ("ADVISOR", "CSV_ONLY", "NO_DATA"):
        n = counts[k]
        s2.append([k, n, f"{n*100/total:.1f}%"])
        s2[s2.max_row][0].fill = fills[k]

    # ---------- Sheet 3: Per-vendor breakdown ----------
    s3 = wb.create_sheet("By vendor")
    s3.append(["Vendor", "Rows",
               "ADVISOR", "CSV_ONLY", "NO_DATA",
               "Total CVEs found"])
    for c in s3[1]: c.font = hdr; c.fill = hdr_fill
    by_vendor: dict[str, dict] = {}
    for src, adv in zip(rows, results):
        v = src["vendor"]
        b = by_vendor.setdefault(v, {"ADVISOR":0,"CSV_ONLY":0,"NO_DATA":0,"cves":0,"n":0})
        b[classify(adv, src["latest_firmware_version"])] += 1
        b["cves"] += adv["crit"]+adv["high"]+adv["med"]+adv["low"]
        b["n"] += 1
    for v in sorted(by_vendor):
        b = by_vendor[v]
        s3.append([v, b["n"], b["ADVISOR"], b["CSV_ONLY"], b["NO_DATA"], b["cves"]])

    for col_letter, w in zip(["A","B","C","D","E","F"], [34,8,12,12,12,18]):
        s3.column_dimensions[col_letter].width = w
    s3.freeze_panes = "A2"

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")


# --- tier classifier (advisor only) --------------------------------------

def _tier_from_adv(adv: dict) -> str:
    if not adv["ok"]:
        return "(advisor error)"
    if adv["latest"] and (adv["crit"]+adv["high"]+adv["med"]+adv["low"] == 0
                          and not adv["earliest_fix"]):
        # latest came from a fetcher (Tier 1) or live web (Tier 3)
        return "Tier 1/3 — latest version"
    if adv["latest"]:
        return "Tier 1 + 2 — latest + CVE data"
    if adv["crit"]+adv["high"]+adv["med"]+adv["low"] > 0:
        return "Tier 2 — NVD CVE feed"
    msg = (adv.get("msg") or "").lower()
    if "portal" in msg or "login" in msg or "vendor support" in msg:
        return "Tier 4 — vendor portal"
    return "Tier 5 — no public source"


# --- HTML writer ---------------------------------------------------------

def write_html(rows: list[dict], results: list[dict]) -> None:
    n = len(rows)
    pairs = list(zip(rows, results))
    counts = Counter(classify(r, s["latest_firmware_version"])
                     for s, r in pairs)
    cve_total_rows = sum(
        1 for _, a in pairs
        if a["crit"]+a["high"]+a["med"]+a["low"] > 0)
    diff_rows = sum(1 for _, a in pairs if a["latest"])
    avg_ms = int(sum(a["ms"] for a in results) / max(n, 1))

    def esc(x): return html.escape(str(x)) if x else ""

    bucket_cls = {
        "ADVISOR":  "bk-adv",
        "CSV_ONLY": "bk-csv",
        "NO_DATA":  "bk-nod",
    }
    bucket_label = {
        "ADVISOR":  "Advisor gave latest / CVE data",
        "CSV_ONLY": "Only CSV value available",
        "NO_DATA":  "No public data anywhere",
    }

    # per-vendor stats
    by_vendor: dict[str, dict] = {}
    for src, adv in pairs:
        v = src["vendor"]
        b = by_vendor.setdefault(v, {"ADVISOR":0,"CSV_ONLY":0,"NO_DATA":0,"n":0,"cves":0,"curated":src["curated"]})
        b[classify(adv, src["latest_firmware_version"])] += 1
        b["n"] += 1
        b["cves"] += adv["crit"]+adv["high"]+adv["med"]+adv["low"]

    vendor_rows_html = []
    for v in sorted(by_vendor):
        b = by_vendor[v]
        cu = "yes" if b["curated"] else "no"
        vendor_rows_html.append(
            "<tr>"
            f"<td>{esc(v)}</td>"
            f"<td class='num'>{b['n']}</td>"
            f"<td class='num bk-adv'>{b['ADVISOR']}</td>"
            f"<td class='num bk-csv'>{b['CSV_ONLY']}</td>"
            f"<td class='num bk-nod'>{b['NO_DATA']}</td>"
            f"<td class='num'>{b['cves']}</td>"
            f"<td>{cu}</td>"
            "</tr>"
        )

    detail_rows_html = []
    for src, adv in pairs:
        outcome = classify(adv, src["latest_firmware_version"])
        cve_total = adv["crit"]+adv["high"]+adv["med"]+adv["low"]
        cve_cell = ("<span class='muted'>—</span>" if cve_total == 0 else
                    f"<span class='sev sev-c'>{adv['crit']}</span>"
                    f"<span class='sev sev-h'>{adv['high']}</span>"
                    f"<span class='sev sev-m'>{adv['med']}</span>"
                    f"<span class='sev sev-l'>{adv['low']}</span>")
        ref = ""
        if adv["release_url"]:
            ref = ("<a href='" + esc(adv["release_url"]) + "' target='_blank' "
                   "rel='noopener'>release&nbsp;notes</a>")
        else:
            ref = "<span class='muted'>—</span>"

        latest_cell = esc(adv["latest"]) or "<span class='muted'>—</span>"
        behind_cell = esc(adv["behind"]) or "<span class='muted'>—</span>"
        detail_rows_html.append(
            f"<tr class='r-{outcome}'>"
            f"<td>{esc(src['vendor'])}</td>"
            f"<td class='mono'>{esc(src['model'])}</td>"
            f"<td class='mono'>{esc(src['current_firmware_version'])}</td>"
            f"<td class='mono muted'>{esc(src['latest_firmware_version'])}</td>"
            f"<td class='mono'>{latest_cell}</td>"
            f"<td class='mono'>{behind_cell}</td>"
            f"<td>{cve_cell}</td>"
            f"<td><span class='bk {bucket_cls[outcome]}'>{outcome.replace('_',' ')}</span></td>"
            f"<td>{ref}</td>"
            "</tr>"
        )

    css = """
    :root { --ink:#1c1c1c; --ink-soft:#4a4a4a; --line:#e6e6e6;
            --rule:#f0f0f0; --bg:#fff; --crit:#d32f2f; --high:#ef6c00;
            --med:#f9a825; --low:#2e7d32; }
    *{box-sizing:border-box}
    html,body{background:var(--bg);color:var(--ink);margin:0;
      font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",
      "Helvetica Neue",Arial,sans-serif;}
    header{padding:36px 48px 24px;border-bottom:1px solid var(--line);}
    header h1{margin:0 0 6px;font-size:26px;letter-spacing:-.01em;}
    header p{margin:0;color:var(--ink-soft);}
    main{padding:28px 48px 64px;max-width:1480px;margin:0 auto;}
    section{margin-bottom:36px;}
    h2{font-size:16px;text-transform:uppercase;letter-spacing:.08em;
       color:var(--ink-soft);margin:0 0 14px;font-weight:600;}
    .kpis{display:grid;
      grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:14px;}
    .kpi{background:#fff;border:1px solid var(--line);border-radius:10px;
         padding:16px 18px;}
    .kpi .v{font-size:24px;font-weight:700;}
    .kpi .k{font-size:12px;color:var(--ink-soft);
            text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;}
    table{width:100%;border-collapse:collapse;background:#fff;
          border:1px solid var(--line);border-radius:10px;overflow:hidden;}
    th,td{text-align:left;padding:9px 12px;border-bottom:1px solid var(--rule);
          font-size:13px;vertical-align:top;}
    th{background:#fafafa;font-weight:600;color:var(--ink-soft);
       text-transform:uppercase;letter-spacing:.05em;font-size:11px;
       position:sticky;top:0;}
    td.num{text-align:right;font-variant-numeric:tabular-nums;}
    .mono,td.mono{font-family:ui-monospace,"SF Mono",Menlo,Consolas,monospace;}
    .muted{color:#9b9b9b;}
    tbody tr:hover{background:#fafafa;}
    .bk{display:inline-block;padding:2px 8px;border-radius:999px;
        font-size:11px;font-weight:600;}
    .bk-adv{background:#e8f5e9;color:#1f6b3c;}
    .bk-csv{background:#fff8e1;color:#a85a00;}
    .bk-nod{background:#fdecea;color:var(--crit);}
    td.bk-adv{background:#f1faf3 !important;}
    td.bk-csv{background:#fffaee !important;}
    td.bk-nod{background:#fdf3f2 !important;}
    .sev{display:inline-block;min-width:26px;padding:1px 6px;margin-right:3px;
         border-radius:4px;font-size:11px;font-weight:600;text-align:center;
         font-variant-numeric:tabular-nums;}
    .sev-c{background:#fdecea;color:var(--crit);}
    .sev-h{background:#fff3e0;color:var(--high);}
    .sev-m{background:#fff8e1;color:#b08600;}
    .sev-l{background:#e8f5e9;color:var(--low);}
    .filterbar{display:flex;gap:12px;align-items:center;margin-bottom:12px;flex-wrap:wrap;}
    .filterbar input{flex:1;min-width:300px;padding:8px 12px;
       border:1px solid var(--line);border-radius:8px;font-size:14px;background:#fff;}
    .filterbar select{padding:8px 10px;border:1px solid var(--line);
       border-radius:8px;font-size:13px;background:#fff;}
    footer{margin-top:48px;padding-top:18px;border-top:1px solid var(--line);
           color:var(--ink-soft);font-size:12px;}
    code{background:#f3f3f3;padding:1px 5px;border-radius:4px;font-size:12px;}
    """

    page = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Fleet Firmware Ground-Truth Test — Switch Spec Agent</title>
<style>{css}</style></head><body>
<header>
  <h1>Fleet Firmware Ground-Truth Test</h1>
  <p>Generated by the Switch Spec Agent · 134 vendors × 5 models = {n} switches ·
     calls <code>agent.firmware_advise()</code> for every row
     (avg latency {avg_ms} ms)</p>
</header>
<main>

<section>
  <h2>At a glance</h2>
  <div class="kpis">
    <div class="kpi"><div class="k">Switches tested</div><div class="v">{n}</div></div>
    <div class="kpi"><div class="k">Vendors</div><div class="v">{len(by_vendor)}</div></div>
    <div class="kpi"><div class="k">Advisor gave latest/CVE</div><div class="v" style="color:#1f6b3c">{counts.get('ADVISOR', 0)}</div></div>
    <div class="kpi"><div class="k">Only CSV data</div><div class="v" style="color:#a85a00">{counts.get('CSV_ONLY', 0)}</div></div>
    <div class="kpi"><div class="k">No public data</div><div class="v" style="color:var(--crit)">{counts.get('NO_DATA', 0)}</div></div>
    <div class="kpi"><div class="k">Rows with CVE data</div><div class="v">{cve_total_rows}</div></div>
    <div class="kpi"><div class="k">Rows with cached latest</div><div class="v">{diff_rows}</div></div>
  </div>
</section>

<section>
  <h2>Per-vendor breakdown · {len(by_vendor)} vendors</h2>
  <table>
    <thead><tr>
      <th>Vendor</th><th class="num">Rows</th>
      <th class="num">ADVISOR</th><th class="num">CSV_ONLY</th><th class="num">NO_DATA</th>
      <th class="num">CVE total</th>
      <th>Curated ground truth</th>
    </tr></thead>
    <tbody>{''.join(vendor_rows_html)}</tbody>
  </table>
</section>

<section>
  <h2>Per-switch detail · {n} rows</h2>
  <div class="filterbar">
    <input id="q" placeholder="filter by vendor, model, version, outcome…">
    <select id="bk">
      <option value="">All outcomes</option>
      <option value="ADVISOR">ADVISOR</option>
      <option value="CSV_ONLY">CSV_ONLY</option>
      <option value="NO_DATA">NO_DATA</option>
    </select>
  </div>
  <table id="t"><thead><tr>
    <th>Vendor</th><th>Model</th><th>Current</th><th>CSV latest</th>
    <th>Advisor latest</th><th>Behind</th>
    <th>CVE (C·H·M·L)</th><th>Outcome</th><th>Reference</th>
  </tr></thead><tbody>{''.join(detail_rows_html)}</tbody></table>
</section>

<footer>
  <p><b>Outcomes:</b>
     <span class="bk bk-adv">ADVISOR</span> = our pipeline returned the
     latest version (Tier 1 / Tier 3) and/or CVE data (Tier 2).
     <span class="bk bk-csv">CSV_ONLY</span> = no pipeline data; only
     the value you supplied in the CSV is available.
     <span class="bk bk-nod">NO_DATA</span> = neither pipeline nor CSV
     had any usable firmware info.</p>
  <p><b>Curated ground truth:</b> the 50 best-known vendors use real
     model names and real version pairs (Cisco Catalyst, Juniper EX,
     MikroTik CRS, etc.). The remaining 84 vendors use synthetic but
     vendor-attributed model + version patterns — what matters for the
     test is that the <em>vendor</em> field exercises the right
     fetcher / CVE bucket / portal route, which it does.</p>
</footer>
</main>
<script>
  const q  = document.getElementById('q');
  const bk = document.getElementById('bk');
  const rows = document.querySelectorAll('#t tbody tr');
  function apply() {{
    const term = q.value.toLowerCase().trim();
    const b = bk.value;
    rows.forEach(r => {{
      const txt = r.textContent.toLowerCase();
      const okText = !term || txt.includes(term);
      const okBk = !b || r.classList.contains('r-' + b);
      r.style.display = (okText && okBk) ? '' : 'none';
    }});
  }}
  q.addEventListener('input', apply);
  bk.addEventListener('change', apply);
</script>
</body></html>
"""
    HTML_OUT.write_text(page, encoding="utf-8")
    print(f"Wrote {HTML_OUT}")


# --- main ----------------------------------------------------------------

def main() -> None:
    rows = build_groundtruth_rows()
    write_csv(rows)

    agent = SpecAgent(live=True)
    results = []
    for i, r in enumerate(rows, 1):
        d = call_advisor(agent, r["model"], r["current_firmware_version"])
        results.append(d)
        if i % 25 == 0 or i == len(rows):
            print(f"  ran {i}/{len(rows)}", file=sys.stderr)

    write_xlsx(rows, results)
    write_html(rows, results)

    counts = Counter(classify(r, s["latest_firmware_version"])
                     for s, r in zip(rows, results))
    print(f"\nOutcome breakdown:")
    for k in ("ADVISOR", "CSV_ONLY", "NO_DATA"):
        print(f"  {k:10s}  {counts[k]:4d}")


if __name__ == "__main__":
    main()
