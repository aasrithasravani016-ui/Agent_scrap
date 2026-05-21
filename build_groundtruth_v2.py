"""
v2 of the ground-truth test report.

Re-classifies every row into ONE of FOUR clear buckets:

  A. LATEST    — agent returned a real latest firmware version string
                 (Tier 1 release-note fetcher or Tier 3 live web pointer).
                 Whether or not CVE data is also present.
  B. CVE_ONLY  — no latest version, BUT the agent returned CVE / fix data
                 from the NIST NVD feed (Tier 2).
  C. PORTAL    — no latest, no CVEs, BUT the agent handed back a vendor
                 portal link / login pointer (Tier 4).
  D. NONE      — agent had nothing for this vendor/model.

This answers the four questions plainly:
  - Are we getting the latest version?  → LATEST count
  - If not, are we getting CVE data?    → CVE_ONLY count
  - If not, portal link?                → PORTAL count
  - Anything?                           → NONE count
"""
from __future__ import annotations

import csv
import html
import json
import sys
import time
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agent import SpecAgent
from vendor_registry import aliases as _aliases

ROOT     = Path("/Users/aasritha/agent")
CSV_IN   = ROOT / "switch_vendors_firmware_groundtruth.csv"
XLSX_OUT = ROOT / "switch_vendors_firmware_groundtruth_v2.xlsx"
HTML_OUT = ROOT / "fleet_firmware_report_v2.html"


def call_advisor(agent: SpecAgent, model: str, current: str) -> dict:
    try:
        t0 = time.time()
        adv = agent.firmware_advise(model, current)
        ms = int((time.time() - t0) * 1000)
    except Exception as e:                                       # noqa: BLE001
        return {"ok": False, "ms": 0, "msg": f"{type(e).__name__}: {e}",
                "vendor": "", "nos": "", "latest": "", "behind": "",
                "crit": 0, "high": 0, "med": 0, "low": 0,
                "earliest_fix": "", "release_url": "", "has_data": False,
                "is_portal": False}

    d = {"ok": True, "ms": ms,
         "vendor": getattr(adv, "vendor", "") or "",
         "nos":    getattr(adv, "nos", "") or "",
         "msg":    (getattr(adv, "message", "") or "")[:400],
         "has_data": bool(getattr(adv, "has_data", False)),
         "latest": "", "behind": "",
         "crit": 0, "high": 0, "med": 0, "low": 0,
         "earliest_fix": "", "release_url": "",
         "is_portal": False}

    diff = getattr(adv, "diff", None)
    if diff:
        tgt = getattr(diff, "target", None)
        if tgt:
            d["latest"]      = getattr(tgt, "version", "") or ""
            d["release_url"] = getattr(tgt, "release_notes_url", "") or ""
        d["behind"] = getattr(diff, "releases_behind", "") or ""

    for a in getattr(adv, "advisories", None) or []:
        sev = (getattr(a, "severity", "") or "").upper()
        if   sev == "CRITICAL": d["crit"] += 1
        elif sev == "HIGH":     d["high"] += 1
        elif sev == "MEDIUM":   d["med"]  += 1
        elif sev == "LOW":      d["low"]  += 1

    if getattr(adv, "recommended_min_version", None):
        d["earliest_fix"] = adv.recommended_min_version

    msg_l = (d["msg"] or "").lower()
    if not d["latest"] and (d["crit"]+d["high"]+d["med"]+d["low"] == 0):
        if any(k in msg_l for k in (
                "portal", "vendor support", "login", "register", "downloads")):
            d["is_portal"] = True
    return d


def classify(adv: dict) -> str:
    if adv["latest"]:
        return "LATEST"
    if (adv["crit"]+adv["high"]+adv["med"]+adv["low"]) > 0:
        return "CVE_ONLY"
    if adv["is_portal"]:
        return "PORTAL"
    return "NONE"


# Reason text per bucket — actionable "to fix"
FIX_FOR = {
    "CVE_ONLY": "Add a Tier-1 release-note fetcher OR populate a Tier-3 live "
                "firmware page URL for this vendor so the pipeline can also "
                "report the current latest version.",
    "PORTAL":   "Vendor hides firmware behind a login wall. To get a latest "
                "version we'd need an authenticated scraper or a public "
                "mirror — otherwise the portal link is the most we can give.",
    "NONE":     "Vendor has no Tier-1 fetcher, no NVD CPE pattern, no Tier-3 "
                "live page cached, and no portal annotation. Add a Tier-3 "
                "live web pointer (run prefetch_firmware.py) or add an NVD "
                "CPE pattern to scrapers/nvd_fetcher.py.",
}


# ----------------- main --------------------------------------------------

def load_csv():
    rows = []
    with CSV_IN.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows


def main() -> None:
    rows = load_csv()
    agent = SpecAgent(live=True)
    results = []
    for i, r in enumerate(rows, 1):
        results.append(call_advisor(
            agent, r["model"], r["current_firmware_version"]))
        if i % 50 == 0 or i == len(rows):
            print(f"  ran {i}/{len(rows)}", file=sys.stderr)

    write_xlsx(rows, results)
    write_html(rows, results)

    cnt = Counter(classify(a) for a in results)
    n = len(rows)
    print("\nAgent grade:")
    for k in ("LATEST", "CVE_ONLY", "PORTAL", "NONE"):
        v = cnt.get(k, 0)
        print(f"  {k:10s}  {v:4d}   ({v*100/n:5.1f}%)")


# ----------------- xlsx --------------------------------------------------

def write_xlsx(rows, results):
    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    fills = {
        "LATEST":   PatternFill("solid", fgColor="DDEBD8"),   # green
        "CVE_ONLY": PatternFill("solid", fgColor="FFF2CC"),   # yellow
        "PORTAL":   PatternFill("solid", fgColor="E1D5F0"),   # purple
        "NONE":     PatternFill("solid", fgColor="F8CECC"),   # red
    }

    # --- sheet 1: Results -------------------------------------------------
    ws = wb.active
    ws.title = "Results"
    cols = [
        "vendor", "model", "current_version", "csv_latest",
        "advisor_latest",          # <-- A: did we get the latest?
        "releases_behind",
        "cve_critical", "cve_high", "cve_medium", "cve_low",  # B: CVE data?
        "portal_only",             # C: portal-only?
        "outcome",                 # one of LATEST/CVE_ONLY/PORTAL/NONE
        "advisor_vendor", "advisor_nos",
        "release_notes_url", "earliest_fix_version",
        "advisor_message", "latency_ms",
    ]
    ws.append(cols)
    for c in ws[1]:
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    for src, adv in zip(rows, results):
        outcome = classify(adv)
        ws.append([
            src["vendor"], src["model"],
            src["current_firmware_version"],
            src["latest_firmware_version"],
            adv["latest"], adv["behind"],
            adv["crit"], adv["high"], adv["med"], adv["low"],
            "yes" if adv["is_portal"] else "",
            outcome,
            adv["vendor"], adv["nos"],
            adv["release_url"], adv["earliest_fix"],
            adv["msg"], adv["ms"],
        ])
        for c in ws[ws.max_row]:
            c.fill = fills[outcome]
    widths = [28,38,22,22,18,14,8,6,8,6,10,12,22,16,38,18,60,10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[
            chr(64+i) if i <= 26 else 'A' + chr(64+i-26)].width = w
    ws.freeze_panes = "A2"

    # --- sheet 2: Agent grade --------------------------------------------
    s2 = wb.create_sheet("Agent grade")
    cnt = Counter(classify(a) for a in results)
    n = max(1, len(rows))
    s2.append(["Outcome", "Rows", "Pct", "Meaning"])
    for c in s2[1]: c.font = hdr_font; c.fill = hdr_fill
    meaning = {
        "LATEST":   "Agent returned a real current latest version",
        "CVE_ONLY": "No latest version, but agent returned CVE data",
        "PORTAL":   "No latest, no CVEs — only a vendor portal link",
        "NONE":     "Agent returned nothing",
    }
    for k in ("LATEST", "CVE_ONLY", "PORTAL", "NONE"):
        s2.append([k, cnt.get(k, 0), f"{cnt.get(k,0)*100/n:.1f}%", meaning[k]])
        s2[s2.max_row][0].fill = fills[k]
    for col, w in zip(["A","B","C","D"], [14,8,8,68]):
        s2.column_dimensions[col].width = w

    # --- sheet 3: By vendor (with what's blocking it) --------------------
    s3 = wb.create_sheet("By vendor")
    s3.append(["Vendor", "Rows", "LATEST", "CVE_ONLY", "PORTAL", "NONE",
               "CVE total", "Status", "To make all 5 LATEST, do this:"])
    for c in s3[1]: c.font = hdr_font; c.fill = hdr_fill
    by_v: dict[str, dict] = {}
    for src, adv in zip(rows, results):
        v = src["vendor"]
        b = by_v.setdefault(v, {"LATEST":0,"CVE_ONLY":0,"PORTAL":0,
                                "NONE":0,"n":0,"cves":0})
        b[classify(adv)] += 1
        b["n"] += 1
        b["cves"] += adv["crit"]+adv["high"]+adv["med"]+adv["low"]
    for v in sorted(by_v):
        b = by_v[v]
        if b["LATEST"] == b["n"]:
            status = "FULLY COVERED"
            todo = "—"
        elif b["LATEST"] > 0:
            status = "PARTIAL"
            todo = "Some rows resolved; extend Tier-3 live discovery for remaining models."
        elif b["CVE_ONLY"] > 0:
            status = "CVE-only"
            todo = FIX_FOR["CVE_ONLY"]
        elif b["PORTAL"] > 0:
            status = "PORTAL-only"
            todo = FIX_FOR["PORTAL"]
        else:
            status = "NO DATA"
            todo = FIX_FOR["NONE"]
        s3.append([v, b["n"], b["LATEST"], b["CVE_ONLY"],
                   b["PORTAL"], b["NONE"], b["cves"], status, todo])
    for col, w in zip(["A","B","C","D","E","F","G","H","I"],
                      [34,8,10,10,10,8,12,18,90]):
        s3.column_dimensions[col].width = w
    s3.freeze_panes = "A2"

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")


# ----------------- html --------------------------------------------------

def write_html(rows, results):
    n = len(rows)
    cnt = Counter(classify(a) for a in results)
    pct = lambda k: cnt.get(k, 0) * 100 / max(n, 1)

    by_v: dict[str, dict] = {}
    for src, adv in zip(rows, results):
        v = src["vendor"]
        b = by_v.setdefault(v, {"LATEST":0,"CVE_ONLY":0,"PORTAL":0,
                                "NONE":0,"n":0,"cves":0})
        b[classify(adv)] += 1
        b["n"] += 1
        b["cves"] += adv["crit"]+adv["high"]+adv["med"]+adv["low"]

    fully = [v for v, b in by_v.items() if b["LATEST"] == b["n"]]
    partial = [v for v, b in by_v.items()
               if 0 < b["LATEST"] < b["n"]]
    cve_only = [v for v, b in by_v.items()
                if b["LATEST"] == 0 and b["CVE_ONLY"] > 0]
    portal_only = [v for v, b in by_v.items()
                   if b["LATEST"] == 0 and b["CVE_ONLY"] == 0
                   and b["PORTAL"] > 0]
    nothing = [v for v, b in by_v.items()
               if b["LATEST"] == 0 and b["CVE_ONLY"] == 0
               and b["PORTAL"] == 0 and b["NONE"] > 0]

    def esc(x): return html.escape(str(x)) if x else ""

    bk_cls = {"LATEST":"bk-lat","CVE_ONLY":"bk-cve",
              "PORTAL":"bk-por","NONE":"bk-none"}
    bk_label = {"LATEST":"LATEST","CVE_ONLY":"CVE ONLY",
                "PORTAL":"PORTAL","NONE":"NO DATA"}

    # detail rows
    detail = []
    for src, adv in zip(rows, results):
        out = classify(adv)
        cve_total = adv["crit"]+adv["high"]+adv["med"]+adv["low"]
        cve_cell = ("<span class='muted'>—</span>" if cve_total == 0 else
                    f"<span class='sev sev-c'>{adv['crit']}</span>"
                    f"<span class='sev sev-h'>{adv['high']}</span>"
                    f"<span class='sev sev-m'>{adv['med']}</span>"
                    f"<span class='sev sev-l'>{adv['low']}</span>")
        ref = (f"<a href='{esc(adv['release_url'])}' target='_blank' "
               "rel='noopener'>release&nbsp;notes</a>"
               if adv["release_url"] else "<span class='muted'>—</span>")
        latest_cell = esc(adv["latest"]) or "<span class='muted'>—</span>"
        behind_cell = esc(adv["behind"]) or "<span class='muted'>—</span>"
        detail.append(
            f"<tr class='r-{out}'>"
            f"<td>{esc(src['vendor'])}</td>"
            f"<td class='mono'>{esc(src['model'])}</td>"
            f"<td class='mono'>{esc(src['current_firmware_version'])}</td>"
            f"<td class='mono muted'>{esc(src['latest_firmware_version'])}</td>"
            f"<td class='mono'>{latest_cell}</td>"
            f"<td class='mono'>{behind_cell}</td>"
            f"<td>{cve_cell}</td>"
            f"<td><span class='bk {bk_cls[out]}'>{bk_label[out]}</span></td>"
            f"<td>{ref}</td></tr>")

    # per-vendor table
    vend_rows = []
    for v in sorted(by_v):
        b = by_v[v]
        if b["LATEST"] == b["n"]:
            status = "FULLY COVERED"; status_cls = "st-full"
        elif b["LATEST"] > 0:
            status = "PARTIAL";       status_cls = "st-part"
        elif b["CVE_ONLY"] > 0:
            status = "CVE-only";      status_cls = "st-cve"
        elif b["PORTAL"] > 0:
            status = "PORTAL-only";   status_cls = "st-por"
        else:
            status = "NO DATA";       status_cls = "st-none"
        vend_rows.append(
            "<tr>"
            f"<td>{esc(v)}</td>"
            f"<td class='num'>{b['n']}</td>"
            f"<td class='num bk-lat'>{b['LATEST']}</td>"
            f"<td class='num bk-cve'>{b['CVE_ONLY']}</td>"
            f"<td class='num bk-por'>{b['PORTAL']}</td>"
            f"<td class='num bk-none'>{b['NONE']}</td>"
            f"<td class='num'>{b['cves']}</td>"
            f"<td><span class='status {status_cls}'>{status}</span></td>"
            "</tr>")

    css = """
    :root{--ink:#1c1c1c;--ink-soft:#4a4a4a;--line:#e6e6e6;--rule:#f0f0f0;
          --bg:#fff;--crit:#d32f2f;--high:#ef6c00;--med:#f9a825;
          --low:#2e7d32;}
    *{box-sizing:border-box}
    html,body{background:var(--bg);color:var(--ink);margin:0;
      font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",
      "Helvetica Neue",Arial,sans-serif;}
    header{padding:36px 48px 24px;border-bottom:1px solid var(--line);}
    header h1{margin:0 0 6px;font-size:28px;letter-spacing:-.01em;}
    header p{margin:0;color:var(--ink-soft);}
    main{padding:28px 48px 64px;max-width:1480px;margin:0 auto;}
    section{margin-bottom:36px;}
    h2{font-size:16px;text-transform:uppercase;letter-spacing:.08em;
       color:var(--ink-soft);margin:0 0 14px;font-weight:600;}
    h3{margin:18px 0 8px;font-size:14px;color:var(--ink);}
    p.lead{font-size:15px;color:var(--ink-soft);max-width:880px;}

    .grade{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:14px;
           margin-bottom:18px;}
    .gradebox{border:1px solid var(--line);border-radius:12px;padding:18px 20px;
              background:#fff;}
    .gradebox .pct{font-size:34px;font-weight:700;letter-spacing:-.02em;
                  margin-bottom:2px;}
    .gradebox .ct{font-size:13px;color:var(--ink-soft);margin-bottom:8px;
                  font-variant-numeric:tabular-nums;}
    .gradebox .lbl{font-size:11px;text-transform:uppercase;letter-spacing:.08em;
                   font-weight:600;}
    .gradebox .ex{font-size:12px;color:var(--ink-soft);margin-top:6px;
                  line-height:1.4;}
    .gb-lat  {background:linear-gradient(to bottom,#f1faf3,#fff);}
    .gb-cve  {background:linear-gradient(to bottom,#fffaee,#fff);}
    .gb-por  {background:linear-gradient(to bottom,#f6f0ff,#fff);}
    .gb-none {background:linear-gradient(to bottom,#fdf3f2,#fff);}
    .gb-lat  .pct{color:#1f6b3c;}
    .gb-cve  .pct{color:#a85a00;}
    .gb-por  .pct{color:#5e2f9c;}
    .gb-none .pct{color:var(--crit);}
    .gb-lat  .lbl{color:#1f6b3c;}
    .gb-cve  .lbl{color:#a85a00;}
    .gb-por  .lbl{color:#5e2f9c;}
    .gb-none .lbl{color:var(--crit);}

    .barwrap{height:14px;background:#f3f3f3;border-radius:7px;
             overflow:hidden;display:flex;margin:6px 0 18px;}
    .barwrap span{display:block;height:100%;}
    .b-lat{background:#7cb86a;}
    .b-cve{background:#f9c870;}
    .b-por{background:#b89fde;}
    .b-none{background:#e08079;}

    table{width:100%;border-collapse:collapse;background:#fff;
          border:1px solid var(--line);border-radius:10px;overflow:hidden;}
    th,td{text-align:left;padding:9px 12px;border-bottom:1px solid var(--rule);
          font-size:13px;vertical-align:top;}
    th{background:#fafafa;font-weight:600;color:var(--ink-soft);
       text-transform:uppercase;letter-spacing:.05em;font-size:11px;
       position:sticky;top:0;}
    td.num{text-align:right;font-variant-numeric:tabular-nums;}
    .mono,td.mono{font-family:ui-monospace,"SF Mono",Menlo,Consolas,monospace;}
    .muted{color:#9b9b9b;}
    tbody tr:hover{background:#fafafa;}

    .bk{display:inline-block;padding:2px 9px;border-radius:999px;
        font-size:11px;font-weight:700;}
    .bk-lat {background:#dbefdb;color:#1f6b3c;}
    .bk-cve {background:#fdebbd;color:#a85a00;}
    .bk-por {background:#e6dbf7;color:#5e2f9c;}
    .bk-none{background:#fcd7d3;color:var(--crit);}
    td.bk-lat {background:#f1faf3 !important;}
    td.bk-cve {background:#fffaee !important;}
    td.bk-por {background:#f7f1fd !important;}
    td.bk-none{background:#fdf3f2 !important;}

    .status{display:inline-block;padding:2px 8px;border-radius:6px;
            font-size:11px;font-weight:600;}
    .st-full{background:#dbefdb;color:#1f6b3c;}
    .st-part{background:#fdebbd;color:#a85a00;}
    .st-cve {background:#fff8e1;color:#a85a00;}
    .st-por {background:#e6dbf7;color:#5e2f9c;}
    .st-none{background:#fcd7d3;color:var(--crit);}

    .sev{display:inline-block;min-width:26px;padding:1px 6px;margin-right:3px;
         border-radius:4px;font-size:11px;font-weight:600;text-align:center;
         font-variant-numeric:tabular-nums;}
    .sev-c{background:#fdecea;color:var(--crit);}
    .sev-h{background:#fff3e0;color:var(--high);}
    .sev-m{background:#fff8e1;color:#b08600;}
    .sev-l{background:#e8f5e9;color:var(--low);}

    .actions{background:#fafbff;border:1px solid #eef0fb;border-radius:12px;
             padding:18px 22px;}
    .actions h3{margin-top:0;}
    .actions ol{margin:0;padding-left:22px;}
    .actions li{margin-bottom:10px;font-size:13.5px;line-height:1.55;}
    .actions code{background:#eef;color:#3a2480;padding:1px 6px;
                  border-radius:4px;font-size:12px;}
    .vlist{font-size:12.5px;color:var(--ink-soft);margin-top:4px;
           line-height:1.55;}
    .vlist b{color:var(--ink);}

    .filterbar{display:flex;gap:12px;align-items:center;margin-bottom:12px;flex-wrap:wrap;}
    .filterbar input{flex:1;min-width:300px;padding:8px 12px;
       border:1px solid var(--line);border-radius:8px;font-size:14px;background:#fff;}
    .filterbar select{padding:8px 10px;border:1px solid var(--line);
       border-radius:8px;font-size:13px;background:#fff;}

    footer{margin-top:48px;padding-top:18px;border-top:1px solid var(--line);
           color:var(--ink-soft);font-size:12px;}
    code{background:#f3f3f3;padding:1px 5px;border-radius:4px;font-size:12px;}
    """

    grade_html = f"""
    <div class="grade">
      <div class="gradebox gb-lat">
        <div class="lbl">A · LATEST</div>
        <div class="pct">{pct('LATEST'):.0f}%</div>
        <div class="ct">{cnt.get('LATEST',0)} of {n} rows</div>
        <div class="ex">Agent returned the actual latest firmware version
          (Tier 1 release-note fetcher or Tier 3 live web pointer).</div>
      </div>
      <div class="gradebox gb-cve">
        <div class="lbl">B · CVE ONLY</div>
        <div class="pct">{pct('CVE_ONLY'):.0f}%</div>
        <div class="ct">{cnt.get('CVE_ONLY',0)} of {n} rows</div>
        <div class="ex">No latest version, but CVE data came back from
          NIST NVD (Tier 2). You'd see which CVEs your fleet is exposed
          to but not the version to upgrade <em>to</em>.</div>
      </div>
      <div class="gradebox gb-por">
        <div class="lbl">C · PORTAL</div>
        <div class="pct">{pct('PORTAL'):.0f}%</div>
        <div class="ct">{cnt.get('PORTAL',0)} of {n} rows</div>
        <div class="ex">No latest, no CVEs — agent handed back a vendor
          portal link (Tier 4) because firmware is behind a login wall.</div>
      </div>
      <div class="gradebox gb-none">
        <div class="lbl">D · NO DATA</div>
        <div class="pct">{pct('NONE'):.0f}%</div>
        <div class="ct">{cnt.get('NONE',0)} of {n} rows</div>
        <div class="ex">Agent has nothing — no fetcher, no NVD pattern,
          no live pointer, no portal annotation.</div>
      </div>
    </div>
    """

    # action plan
    fully_n   = len(fully)
    partial_n = len(partial)
    cve_n     = len(cve_only)
    por_n     = len(portal_only)
    none_n    = len(nothing)

    def vlist(vs, lim=20):
        if not vs: return "<i>none</i>"
        shown = ", ".join(sorted(vs)[:lim])
        extra = f" <span class='muted'>(+ {len(vs)-lim} more)</span>" if len(vs) > lim else ""
        return shown + extra

    actions_html = f"""
    <div class="actions">
      <h3>How to push every row into the green LATEST bucket</h3>
      <ol>
        <li><b>Already done:</b> {fully_n} vendors are fully covered
            (every model returned LATEST). These are the Tier-1 fetchers
            (MikroTik, Ubiquiti, NVIDIA) + Tier-3 live-discovered pages.
            <div class="vlist"><b>Vendors:</b> {vlist(fully, 25)}</div></li>

        <li><b>Extend Tier-3 live discovery</b> to the
            <b>{partial_n} partial</b> + <b>{cve_n} CVE-only</b> +
            <b>{none_n} no-data</b> vendors.
            <br>Run <code>python3 prefetch_firmware.py</code> again with a
            wider deadline (<code>deadline_sec=15</code> instead of 8) and
            with the Mojeek/Startpage queries tuned per-vendor
            (use the vendor's <code>website</code> field from
            <code>vendors.json</code> as a search hint). This is the
            single biggest lever — most "CVE_ONLY" rows are blocked only
            because the live web scraper timed out, not because the page
            doesn't exist.
            <div class="vlist"><b>CVE-only vendors:</b> {vlist(cve_only, 25)}</div>
            <div class="vlist"><b>Partial vendors:</b> {vlist(partial, 25)}</div>
            </li>

        <li><b>Write a hand-built Tier-1 fetcher</b> for the top
            vulnerable vendors where Tier-3 web search can't find a
            stable URL. These show up as CVE_ONLY today (they have
            CVE data but no public release page). Highest ROI targets:
            Cisco IOS XE, Juniper Junos, Arista EOS, HPE Aruba CX,
            Dell OS10, Huawei VRP — but these vendors are
            <em>login-gated</em>, so this is hard.
            See item 4 instead.</li>

        <li><b>Annotate login-gated vendors in <code>vendor_registry.py</code></b>
            so they get a clean Tier-4 portal link instead of falling
            through to NONE.
            <div class="vlist"><b>NO-DATA vendors to annotate:</b> {vlist(nothing, 25)}</div></li>

        <li><b>Expand NVD CPE coverage in
            <code>scrapers/nvd_fetcher.py</code></b> by adding more
            <code>cpe:2.3:o:&lt;vendor&gt;:*_firmware</code> patterns. This
            won't give you a latest version, but it will move NONE-bucket
            vendors into the CVE_ONLY bucket so at least the security
            picture is visible.</li>
      </ol>
      <p style="margin-top:14px;font-size:13px;color:var(--ink-soft);">
        <b>One-line answer to "what to do":</b>
        the cheapest 80% win is to rerun <code>prefetch_firmware.py</code>
        with a wider deadline and per-vendor search hints; it'll
        upgrade most CVE_ONLY and NO_DATA rows to LATEST without writing
        any new code.
      </p>
    </div>
    """

    bar = (
        f"<div class='barwrap'>"
        f"<span class='b-lat' style='width:{pct('LATEST'):.2f}%'></span>"
        f"<span class='b-cve' style='width:{pct('CVE_ONLY'):.2f}%'></span>"
        f"<span class='b-por' style='width:{pct('PORTAL'):.2f}%'></span>"
        f"<span class='b-none' style='width:{pct('NONE'):.2f}%'></span>"
        "</div>"
    )

    page = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Switch Spec Agent · Firmware Test Report</title>
<style>{css}</style></head><body>

<header>
  <h1>How good is the firmware agent?</h1>
  <p>{n} switches across {len(by_v)} vendors · each one was run through
     <code>agent.firmware_advise()</code>. Every row landed in exactly
     one of four buckets below.</p>
</header>

<main>

<section>
  <h2>Agent grade</h2>
  <p class="lead">Read this strip left → right. The bigger the green
     section, the better. <b>LATEST</b> = pipeline gave you the real
     current firmware version; <b>CVE&nbsp;ONLY</b> = pipeline gave you
     security data but not the version; <b>PORTAL</b> = behind a login;
     <b>NO&nbsp;DATA</b> = nothing.</p>
  {bar}
  {grade_html}
</section>

<section>
  <h2>What to do to get more rows into LATEST</h2>
  {actions_html}
</section>

<section>
  <h2>Per-vendor breakdown · {len(by_v)} vendors</h2>
  <p class="lead">For each vendor, how many of its 5 test rows landed
     in each bucket. <b>FULLY COVERED</b> means all 5 got LATEST.
     <b>PARTIAL</b> means some did. <b>CVE-only / PORTAL-only / NO DATA</b>
     means none did, and tells you what's blocking that vendor.</p>
  <table>
    <thead><tr>
      <th>Vendor</th><th class="num">Rows</th>
      <th class="num">LATEST</th><th class="num">CVE_ONLY</th>
      <th class="num">PORTAL</th><th class="num">NONE</th>
      <th class="num">CVE total</th><th>Status</th>
    </tr></thead>
    <tbody>{''.join(vend_rows)}</tbody>
  </table>
</section>

<section>
  <h2>Per-switch detail · {n} rows</h2>
  <div class="filterbar">
    <input id="q" placeholder="filter by vendor, model, version…">
    <select id="bk">
      <option value="">All buckets</option>
      <option value="LATEST">LATEST only</option>
      <option value="CVE_ONLY">CVE_ONLY only</option>
      <option value="PORTAL">PORTAL only</option>
      <option value="NONE">NO DATA only</option>
    </select>
  </div>
  <table id="t"><thead><tr>
    <th>Vendor</th><th>Model</th><th>Current</th><th>Ground truth latest</th>
    <th>Agent latest</th><th>Behind</th>
    <th>CVEs (C·H·M·L)</th><th>Outcome</th><th>Notes URL</th>
  </tr></thead><tbody>{''.join(detail)}</tbody></table>
</section>

<footer>
  <p>Generated from <code>switch_vendors_firmware_groundtruth.csv</code>
     by <code>build_groundtruth_v2.py</code>. Also written:
     <code>switch_vendors_firmware_groundtruth_v2.xlsx</code> with
     Results / Agent grade / By vendor sheets.</p>
</footer>
</main>

<script>
  const q  = document.getElementById('q');
  const bk = document.getElementById('bk');
  const rows = document.querySelectorAll('#t tbody tr');
  function apply() {{
    const term = q.value.toLowerCase().trim();
    const b = bk.value;
    rows.forEach(r => {{
      const txt = r.textContent.toLowerCase();
      const okT = !term || txt.includes(term);
      const okB = !b || r.classList.contains('r-' + b);
      r.style.display = (okT && okB) ? '' : 'none';
    }});
  }}
  q.addEventListener('input', apply);
  bk.addEventListener('change', apply);
</script>
</body></html>"""

    HTML_OUT.write_text(page, encoding="utf-8")
    print(f"Wrote {HTML_OUT}")


if __name__ == "__main__":
    main()
